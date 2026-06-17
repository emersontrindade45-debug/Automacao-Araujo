# -*- coding: utf-8 -*-
"""
Aplica acoes em massa na tabela `produtos` a partir de uma planilha com
coluna `acao` (upsert, ativar, desativar). Reaproveita o padrao de lotes
pequenos com pausa de dados/atualizar_precos.py.

Acoes aceitas (coluna `acao`, case-insensitive, vazio = upsert):
  upsert     -> cria por nome (se nao existe) ou atualiza todos os campos
                preenchidos na linha (unidade, preco_atual, categoria, nicho,
                validade). Sempre forca ativo=true.
  ativar     -> so altera ativo=true para o nome. Demais colunas ignoradas.
  desativar  -> so altera ativo=false para o nome. Demais colunas ignoradas.

ativar/desativar para um nome que nao existe no banco: linha pulada e
reportada no log final, nao interrompe o processamento.

Uso:
  python aplicar_acoes_produtos.py <arquivo.xlsx>
  python aplicar_acoes_produtos.py <arquivo.xlsx> --dry-run
  python aplicar_acoes_produtos.py <arquivo.xlsx> --limite 100
"""
import json, sys, time, urllib.parse
import openpyxl
import urllib.request, urllib.error

ENV = "../.env.local"
BATCH = 50
PAUSA_SEGUNDOS = 8
ACOES_VALIDAS = {"upsert", "ativar", "desativar", ""}


def carregar_env():
    url = key = None
    for line in open(ENV, encoding="utf-8", errors="replace"):
        line = line.strip()
        if line.startswith("NEXT_PUBLIC_SUPABASE_URL"):
            url = line.split("=", 1)[1].strip().strip('"').strip("'")
        if line.startswith("SUPABASE_SERVICE_ROLE_KEY"):
            key = line.split("=", 1)[1].strip().strip('"').strip("'")
    if not url or not key:
        raise SystemExit("NEXT_PUBLIC_SUPABASE_URL ou SUPABASE_SERVICE_ROLE_KEY nao encontrados em " + ENV)
    return url.rstrip("/"), key


def supabase_request(url, key, method, endpoint_suffix, body, dry_run):
    if dry_run:
        return
    req = urllib.request.Request(
        f"{url}{endpoint_suffix}",
        data=json.dumps(body).encode("utf-8") if body is not None else None,
        method=method,
        headers={
            "apikey": key,
            "Authorization": f"Bearer {key}",
            "Content-Type": "application/json",
            "Prefer": "resolution=merge-duplicates,return=minimal" if method == "POST" else "return=minimal",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as r:
            return
    except urllib.error.HTTPError as e:
        body_err = e.read().decode("utf-8", errors="replace")
        raise SystemExit(f"Erro HTTP {e.code} ({method} {endpoint_suffix}): {body_err}")


def ler_planilha(caminho):
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb["Produtos"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers)}

    linhas_upsert = []
    nomes_ativar = []
    nomes_desativar = []
    erros = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        nome = row[col["nome"]] if "nome" in col else None
        if not nome or not str(nome).strip():
            erros.append(("(vazio)", "nome_vazio"))
            continue
        nome = str(nome).strip()

        acao_raw = row[col["acao"]] if "acao" in col and row[col["acao"]] else ""
        acao = str(acao_raw).strip().lower()
        if acao not in ACOES_VALIDAS:
            erros.append((nome, f"acao_invalida:{acao_raw}"))
            continue

        if acao == "ativar":
            nomes_ativar.append(nome)
        elif acao == "desativar":
            nomes_desativar.append(nome)
        else:  # upsert ou vazio
            preco = row[col["preco_atual"]] if "preco_atual" in col else None
            unidade = row[col["unidade"]] if "unidade" in col and row[col["unidade"]] else "UN"
            categoria = row[col["categoria"]] if "categoria" in col and row[col["categoria"]] else None
            nicho = row[col["nicho"]] if "nicho" in col and row[col["nicho"]] else None
            validade = row[col["validade"]] if "validade" in col and row[col["validade"]] else None
            if preco is None:
                erros.append((nome, "upsert_sem_preco"))
                continue
            linhas_upsert.append({
                "nome": nome,
                "preco_atual": round(float(preco), 2),
                "unidade": str(unidade).strip(),
                "categoria": str(categoria).strip().lower() if categoria else None,
                "nicho": str(nicho).strip().lower() if nicho else None,
                "validade": str(validade).strip() if validade else None,
                "ativo": True,
            })

    wb.close()
    return linhas_upsert, nomes_ativar, nomes_desativar, erros


def verificar_existentes(url, key, nomes, dry_run):
    """Retorna o subconjunto de `nomes` que de fato existe na tabela produtos."""
    if dry_run or not nomes:
        return set(nomes)
    existentes = set()
    for i in range(0, len(nomes), BATCH):
        lote = nomes[i:i + BATCH]
        filtro = ",".join(f'"{n}"' for n in lote)
        filtro_encoded = urllib.parse.quote(filtro)
        req = urllib.request.Request(
            f"{url}/rest/v1/produtos?nome=in.({filtro_encoded})&select=nome",
            method="GET",
            headers={"apikey": key, "Authorization": f"Bearer {key}"},
        )
        with urllib.request.urlopen(req, timeout=60) as r:
            data = json.loads(r.read().decode("utf-8"))
            existentes.update(item["nome"] for item in data)
    return existentes


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    if not args:
        raise SystemExit("Uso: python aplicar_acoes_produtos.py <arquivo.xlsx> [--dry-run] [--limite N]")
    caminho = args[0]

    dry_run = "--dry-run" in sys.argv
    limite = None
    if "--limite" in sys.argv:
        limite = int(sys.argv[sys.argv.index("--limite") + 1])

    if dry_run:
        print("[DRY-RUN] Nenhuma alteracao sera feita no banco.")

    url, key = carregar_env()

    linhas_upsert, nomes_ativar, nomes_desativar, erros = ler_planilha(caminho)

    if limite:
        linhas_upsert = linhas_upsert[:limite]
        nomes_ativar = nomes_ativar[:limite]
        nomes_desativar = nomes_desativar[:limite]

    print(f"upsert: {len(linhas_upsert)} | ativar: {len(nomes_ativar)} | desativar: {len(nomes_desativar)} | erros de leitura: {len(erros)}")

    # valida existencia antes de ativar/desativar
    existentes_ativar = verificar_existentes(url, key, nomes_ativar, dry_run)
    existentes_desativar = verificar_existentes(url, key, nomes_desativar, dry_run)
    nao_encontrados = (set(nomes_ativar) - existentes_ativar) | (set(nomes_desativar) - existentes_desativar)
    for nome in nao_encontrados:
        erros.append((nome, "nome_nao_encontrado_no_banco"))

    nomes_ativar = [n for n in nomes_ativar if n in existentes_ativar]
    nomes_desativar = [n for n in nomes_desativar if n in existentes_desativar]

    total_aplicado = 0

    # upsert em lotes
    for i in range(0, len(linhas_upsert), BATCH):
        lote = linhas_upsert[i:i + BATCH]
        supabase_request(url, key, "POST", "/rest/v1/produtos?on_conflict=nome", lote, dry_run)
        total_aplicado += len(lote)
        print(f"  upsert {min(i+BATCH, len(linhas_upsert))}/{len(linhas_upsert)}", flush=True)
        if i + BATCH < len(linhas_upsert) and not dry_run:
            time.sleep(PAUSA_SEGUNDOS)

    # ativar em lotes (PATCH com filtro nome=in.(...))
    for i in range(0, len(nomes_ativar), BATCH):
        lote = nomes_ativar[i:i + BATCH]
        filtro = ",".join(f'"{n}"' for n in lote)
        filtro_encoded = urllib.parse.quote(filtro)
        supabase_request(url, key, "PATCH", f"/rest/v1/produtos?nome=in.({filtro_encoded})", {"ativo": True}, dry_run)
        total_aplicado += len(lote)
        print(f"  ativar {min(i+BATCH, len(nomes_ativar))}/{len(nomes_ativar)}", flush=True)
        if i + BATCH < len(nomes_ativar) and not dry_run:
            time.sleep(PAUSA_SEGUNDOS)

    # desativar em lotes
    for i in range(0, len(nomes_desativar), BATCH):
        lote = nomes_desativar[i:i + BATCH]
        filtro = ",".join(f'"{n}"' for n in lote)
        filtro_encoded = urllib.parse.quote(filtro)
        supabase_request(url, key, "PATCH", f"/rest/v1/produtos?nome=in.({filtro_encoded})", {"ativo": False}, dry_run)
        total_aplicado += len(lote)
        print(f"  desativar {min(i+BATCH, len(nomes_desativar))}/{len(nomes_desativar)}", flush=True)
        if i + BATCH < len(nomes_desativar) and not dry_run:
            time.sleep(PAUSA_SEGUNDOS)

    print(f"\n{'[DRY-RUN] ' if dry_run else ''}Concluido: {total_aplicado} acoes aplicadas.")
    if erros:
        print(f"\n{len(erros)} linha(s) ignorada(s):")
        for nome, motivo in erros:
            print(f"  {nome}: {motivo}")


if __name__ == "__main__":
    main()
