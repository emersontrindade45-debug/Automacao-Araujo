# -*- coding: utf-8 -*-
"""
Atualiza preco_atual (e ativa) dos produtos ja existentes na tabela `produtos`
a partir do arquivo conciliado produtos_hub_precos_conciliados.xlsx.

- Casa por `nome` (UNIQUE na tabela) via upsert com on_conflict=nome,
  Prefer: resolution=merge-duplicates -> atualiza so as colunas enviadas
  (preco_atual, ativo) na linha existente, sem duplicar nem alterar embedding.
- Cada linha atualizada dispara o trigger produtos_embedding_trigger (webhook
  n8n -> OpenAI -> recalcula embedding). Por isso enviamos em lotes pequenos
  com pausa entre eles, para nao gerar um pico de 27 mil chamadas simultaneas.

Uso:
  python atualizar_precos.py            # atualiza tudo
  python atualizar_precos.py --dry-run  # mostra quantos seriam atualizados, sem tocar no banco
  python atualizar_precos.py --limite 100  # so os N primeiros (teste)
"""
import json, sys, time
import openpyxl
import urllib.request, urllib.error

XLSX = "produtos_hub_precos_conciliados.xlsx"
ENV = "../.env.local"
BATCH = 50           # produtos por lote (cada um dispara 1 webhook de embedding)
PAUSA_SEGUNDOS = 8    # pausa entre lotes para nao saturar n8n/OpenAI


def carregar_env():
    url = key = None
    for line in open(ENV, encoding="utf-8", errors="replace"):
        line = line.strip()
        if line.startswith("NEXT_PUBLIC_SUPABASE_URL"):
            url = line.split("=", 1)[1].strip().strip('"').strip("'")
        if line.startswith("SUPABASE_SERVICE_ROLE_KEY"):
            key = line.split("=", 1)[1].strip().strip('"').strip("'")
    if not url or not key:
        raise SystemExit("NEXT_PUBLIC_SUPABASE_URL ou SUPABASE_SERVICE_ROLE_KEY nao encontrados em " + ENV)
    return url.rstrip("/"), key


def supabase_upsert_precos(url, key, rows, dry_run=False):
    endpoint = f"{url}/rest/v1/produtos?on_conflict=nome"
    body = json.dumps(rows).encode("utf-8")
    if dry_run:
        return len(rows)
    req = urllib.request.Request(
        endpoint,
        data=body,
        method="POST",
        headers={
            "apikey": key,
            "Authorization": f"Bearer {key}",
            "Content-Type": "application/json",
            "Prefer": "resolution=merge-duplicates,return=minimal",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as r:
            return len(rows)
    except urllib.error.HTTPError as e:
        body_err = e.read().decode("utf-8", errors="replace")
        raise SystemExit(f"Erro HTTP {e.code}: {body_err}")


def main():
    dry_run = "--dry-run" in sys.argv
    limite = None
    if "--limite" in sys.argv:
        limite = int(sys.argv[sys.argv.index("--limite") + 1])

    if dry_run:
        print("[DRY-RUN] Nenhuma alteracao sera feita no banco.")

    url, key = carregar_env()

    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["Produtos"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers)}

    # nome e UNIQUE na tabela -- quando o mesmo nome aparece mais de uma vez no
    # xlsx (produtos com codigos ERP diferentes que a IA tratou como nome igual),
    # mantemos o menor preco entre as opcoes.
    por_nome = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        nome = row[col["nome"]]
        preco = row[col["preco_atual"]]
        unidade = row[col["unidade"]] if "unidade" in col and row[col["unidade"]] else "UN"
        if not nome or preco is None:
            continue
        nome = str(nome).strip()
        preco = round(float(preco), 2)
        if nome not in por_nome or preco < por_nome[nome]["preco_atual"]:
            por_nome[nome] = {
                "nome": nome,
                "preco_atual": preco,
                "unidade": str(unidade).strip(),
                "ativo": True,
            }
    wb.close()

    rows_to_update = list(por_nome.values())

    if limite:
        rows_to_update = rows_to_update[:limite]

    print(f"Produtos a atualizar: {len(rows_to_update)}")
    if not rows_to_update:
        print("Nada para atualizar.")
        return

    lotes = [rows_to_update[i:i + BATCH] for i in range(0, len(rows_to_update), BATCH)]
    print(f"Enviando {len(lotes)} lotes de até {BATCH}, pausa de {PAUSA_SEGUNDOS}s entre lotes...")

    total = 0
    t0 = time.time()
    for i, lote in enumerate(lotes, 1):
        n = supabase_upsert_precos(url, key, lote, dry_run=dry_run)
        total += n
        el = time.time() - t0
        print(f"  lote {i}/{len(lotes)} | {total} enviados | {el:.0f}s", flush=True)
        if i < len(lotes) and not dry_run:
            time.sleep(PAUSA_SEGUNDOS)

    print(f"\n{'[DRY-RUN] ' if dry_run else ''}Concluido: {total} produtos {'seriam ' if dry_run else ''}atualizados.")
    print("Cada produto atualizado dispara o trigger de embedding (webhook n8n -> OpenAI) automaticamente.")


if __name__ == "__main__":
    main()
