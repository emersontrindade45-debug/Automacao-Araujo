# -*- coding: utf-8 -*-
"""
Concilia os itens curados do BANCO (Hub) que faltam na planilha tratada do ERP.
Adiciona ofertas, kits e cortes de carne/padaria que existem so no Supabase, SEM duplicar.

Entrada:
  produtos_hub_TRATADO.xlsx   (27.161 itens do ERP, ja tratados)
  banco_faltantes.json        (124 itens do banco ausentes na planilha — gerado no cruzamento)

Regras (decididas com o usuario, 2026-06-24):
  - 31 ofertas (tipo=oferta) -> TRAZER todas (vitrine intencional, coexistem com o corte comum)
  - 11 kits  -> TRAZER (dedup do 'Kit Churrasco' que aparece 2x no banco)
  - 56 carnes + 26 padaria -> TRAZER, exceto:
      * 5 com preco R$0,00 -> DESCARTAR
      * 3 quase-identicos (dup real) -> DESCARTAR: Bolo de Cenoura com Chocolate,
        Bolo de Milho, Sardinha Tomate Gomes da Costa 125g
  - Schema do Hub: nome, unidade, preco_atual, categoria, nicho, validade.
    nicho do banco PRESERVADO (acougue/churrasco/padaria) — esses sao curados.
  - dedup final por nome (case-insensitive) p/ garantir 0 duplicata na planilha resultante.

Saida:
  produtos_hub_CONCILIADO.xlsx   (planilha tratada + faltantes do banco)
  rel_conciliacao_adicionados.csv

Uso: python -u conciliar_banco_hub.py
"""
import csv
import json
import re
import sys
import unicodedata

import openpyxl

PLAN = "produtos_hub_TRATADO.xlsx"
FALT = "banco_faltantes.json"
OUT = "produtos_hub_CONCILIADO.xlsx"
REL = "rel_conciliacao_adicionados.csv"

# descartes explicitos (quase-identicos = dup real) por nome lower
DESCARTAR_NOMES = {
    "bolo de cenoura com chocolate",
    "bolo de milho",
    "sardinha tomate gomes da costa 125g",
}


def sa(s):
    return "".join(c for c in unicodedata.normalize("NFD", str(s or "")) if unicodedata.category(c) != "Mn")


def ck(s):
    return re.sub(r"[^a-z0-9]+", "", sa(s).lower())


def fl(v):
    try:
        return round(float(str(v).replace(",", ".").strip()), 2)
    except (ValueError, TypeError):
        return 0.0


def main():
    # 1. planilha tratada -> linhas + indice de chaves
    wb = openpyxl.load_workbook(PLAN, read_only=True, data_only=True)
    ws = wb["Produtos"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    plan_rows = []
    plan_chaves = set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        plan_rows.append(list(r))
        plan_chaves.add(ck(r[0]))
    wb.close()
    print(f"Planilha tratada: {len(plan_rows)} | colunas: {headers}", flush=True)

    # 2. faltantes do banco
    falt = json.load(open(FALT, encoding="utf-8"))
    print(f"Faltantes do banco: {len(falt)}", flush=True)

    adicionados = []
    descartados = []
    vistos = set(plan_chaves)  # evita duplicar contra a planilha E entre si

    for b in falt:
        nome = str(b["nome"]).strip()
        nlow = nome.lower()
        preco = fl(b["preco_atual"])
        k = ck(nome)

        # descartes
        if preco <= 0:
            descartados.append((nome, preco, "preco_zero"))
            continue
        if nlow in DESCARTAR_NOMES:
            descartados.append((nome, preco, "quase_identico"))
            continue
        if k in vistos:
            descartados.append((nome, preco, "ja_presente_ou_duplicado"))
            continue

        vistos.add(k)
        cat = (b.get("categoria") or "").strip().lower() or None
        nicho = (b.get("nicho") or "").strip().lower() or None
        if nicho in ("none", ""):
            nicho = None
        unidade = (b.get("unidade") or "UN").strip()
        validade = b.get("validade") or None
        if validade in ("None", "", "none"):
            validade = None

        # linha no MESMO formato do header da planilha
        linha = {
            "nome": nome,
            "unidade": unidade,
            "preco_atual": preco,
            "categoria": cat,
            "nicho": nicho,
            "validade": validade,
        }
        plan_rows.append([linha.get(h) for h in headers])
        adicionados.append((nome, preco, cat, nicho, b.get("tipo")))

    print(f"Adicionados: {len(adicionados)} | Descartados: {len(descartados)}", flush=True)

    # 3. gravar planilha conciliada
    wb_out = openpyxl.Workbook()
    wso = wb_out.active
    wso.title = "Produtos"
    wso.append(headers)
    for row in sorted(plan_rows, key=lambda x: (str(x[headers.index("categoria")] or ""), str(x[0]))):
        wso.append(row)
    wb_out.save(OUT)
    print(f"OK -> {OUT} ({len(plan_rows)} linhas)", flush=True)

    # 4. relatorio
    with open(REL, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["nome", "preco", "categoria", "nicho", "tipo_origem_banco"])
        for n, p, c, nh, t in sorted(adicionados, key=lambda x: (x[4] or "", x[0])):
            w.writerow([n, p, c, nh, t])

    # resumo por tipo de origem
    from collections import Counter
    print("\n=== Adicionados por tipo (origem banco) ===", flush=True)
    for t, n in Counter(a[4] for a in adicionados).most_common():
        print(f"  {t}: {n}", flush=True)
    print("\n=== Descartados por motivo ===", flush=True)
    for m, n in Counter(d[2] for d in descartados).most_common():
        print(f"  {m}: {n}", flush=True)


if __name__ == "__main__":
    main()
