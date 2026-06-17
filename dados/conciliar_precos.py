# -*- coding: utf-8 -*-
"""
Concilia precos do ERP com o relatorio padrao do hub.

Cadeia de match (nome do hub nao tem codigo_erp, entao passamos por um meio):
  hub (nome)  --chave_normalizada-->  produtos_tratado_v2.csv (nome + _codigo_erp)
              --_codigo_erp-->        planilha de precos do ERP (codigo + PrVenda)

Dentro de cada chave normalizada que se repete (produtos com nome igual),
o pareamento e feito por posicao relativa dentro do grupo, pois ambos os
arquivos vem do mesmo export do ERP e preservam a ordem original.

Qualquer linha do hub cuja chave nao apareca no tratado_v2, ou cujo grupo
tenha contagem diferente entre os dois arquivos, fica de fora do arquivo
final e vai para um relatorio de pendencias (nao adivinha preco).

Entradas:
  produtos_relatorio_original_formatação_padrão_hub.xlsx  (aba Produtos)
  produtos_tratado_v2.csv
  relatorio de produtos para IA tabela preco de venda.xlsx (aba Plan1)

Saidas:
  produtos_hub_precos_conciliados.xlsx  -> mesmo formato do hub, preco_atual preenchido
  conciliacao_pendencias.csv            -> linhas do hub que NAO puderam ser conciliadas
"""
import csv
import re
import unicodedata
from collections import defaultdict

import openpyxl

HUB_XLSX = "produtos_relatorio_original_formatação_padrão_hub.xlsx"
TRATADO_CSV = "produtos_tratado_v2.csv"
PRECO_XLSX = "relatorio de produtos para IA tabela preco de venda.xlsx"
OUT_XLSX = "produtos_hub_precos_conciliados.xlsx"
OUT_PENDENCIAS = "conciliacao_pendencias.csv"

COL_PRVENDA = 64  # indice 0-based da coluna "PrVenda" na aba Plan1


def chave(s):
    s = str(s or "").lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9]+", "", s)


def palavras(s):
    s = str(s or "").lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return [w for w in s.split() if w]


# qualificadores que podem indicar produto/corte diferente -- nao aceitar por prefixo
PALAVRAS_AMBIGUAS = {
    "recheado", "recheada", "atacado", "resfriada", "resfriado", "congelada", "congelado",
    "fatiada", "fatiado", "temperada", "temperado", "defumada", "defumado",
    "desossada", "desossado", "inteira", "inteiro", "moida", "moido",
    "especial", "premium", "tradicional", "light", "diet", "zero",
}


# overrides manuais: nome no hub ficou com a unidade quebrada na coluna errada (corrupcao
# de leitura do xlsx original), confirmados um a um contra o _nome_erp/_codigo_erp do tratado_v2.
OVERRIDES_CODIGO_ERP = {
    "Creme de Morango": "146445",                              # "Creme de Morango, 2,05kg."
    "Desodorante Rexona Roll-On": "5558",                      # "Desodorante Rexona Roll-On, 50ml, 72h."
    "Inseticida SBP": "17522",                                 # "Inseticida SBP, 300ml, Gotas Suave."
    "Tintura para cabelo da marca Kanechom": "5922",           # "..., 118g, cor Vermelho Cereja."
    "Rodo de Plástico": "116383",                          # "Rodo de Plástico, 60cm, da Casa das Vassouras."
    "Lâmina de Aço Assolan": "140034",                # "..., 10 unidades, 45g."
}


def main():
    # 1. planilha de precos do ERP: codigo -> preco
    wb_preco = openpyxl.load_workbook(PRECO_XLSX, read_only=True, data_only=True)
    ws_preco = wb_preco["Plan1"]
    preco_por_codigo = {}
    for row in ws_preco.iter_rows(min_row=5, values_only=True):
        if row[0] is None:
            continue
        preco_por_codigo[row[0]] = row[COL_PRVENDA]
    wb_preco.close()
    print(f"Precos carregados do ERP: {len(preco_por_codigo)}")

    # 2. tratado_v2: agrupar por chave normalizada, preservando ordem
    with open(TRATADO_CSV, encoding="utf-8-sig") as f:
        tratado_rows = list(csv.DictReader(f))
    grupos_tratado = defaultdict(list)
    for r in tratado_rows:
        grupos_tratado[chave(r["nome"])].append(r["_codigo_erp"])

    # indice auxiliar para match por prefixo de palavras (fallback quando a chave exata nao bate)
    trat_palavras = [palavras(r["nome"]) for r in tratado_rows]

    def match_por_prefixo(nome_hub):
        """Busca candidato unico no tratado cujo nome comeca exatamente com as palavras do hub.
        Recusa se a(s) palavra(s) extra contiver qualificador ambiguo (pode ser produto/corte diferente)."""
        hw = palavras(nome_hub)
        if not hw:
            return None
        candidatos = [i for i, tw in enumerate(trat_palavras) if len(tw) >= len(hw) and tw[: len(hw)] == hw]
        if len(candidatos) != 1:
            return None
        i = candidatos[0]
        extra = trat_palavras[i][len(hw):]
        if any(w in PALAVRAS_AMBIGUAS for w in extra):
            return None
        return tratado_rows[i]["_codigo_erp"]

    # 3. hub: ler todas as linhas, agrupar por chave preservando ordem
    wb_hub = openpyxl.load_workbook(HUB_XLSX, read_only=True, data_only=True)
    ws_hub = wb_hub["Produtos"]
    headers = [c.value for c in next(ws_hub.iter_rows(min_row=1, max_row=1))]
    hub_rows = [list(r) for r in ws_hub.iter_rows(min_row=2, values_only=True) if r[0]]
    wb_hub.close()
    print(f"Linhas do hub: {len(hub_rows)}")

    idx_preco = headers.index("preco_atual")

    grupos_hub = defaultdict(list)
    for i, row in enumerate(hub_rows):
        grupos_hub[chave(row[0])].append(i)

    conciliados = []
    pendencias = []

    for k, idxs_hub in grupos_hub.items():
        idxs_trat = grupos_tratado.get(k, [])
        if len(idxs_trat) != len(idxs_hub):
            # sem match exato (ou contagem divergente) -- tenta prefixo de palavras quando for caso 1-para-0
            if not idxs_trat and len(idxs_hub) == 1:
                nome_hub = hub_rows[idxs_hub[0]][0]
                codigo_erp = OVERRIDES_CODIGO_ERP.get(nome_hub) or match_por_prefixo(nome_hub)
                if codigo_erp is not None:
                    idxs_trat = [codigo_erp]
                else:
                    pendencias.append((hub_rows[idxs_hub[0]][0], "nome_nao_encontrado_no_tratado"))
                    continue
            else:
                for i in idxs_hub:
                    motivo = "nome_nao_encontrado_no_tratado" if not idxs_trat else "contagem_divergente_no_grupo"
                    pendencias.append((hub_rows[i][0], motivo))
                continue
        for pos, i in enumerate(idxs_hub):
            codigo_erp = idxs_trat[pos]
            codigo_erp_int = int(codigo_erp)
            if codigo_erp_int not in preco_por_codigo:
                pendencias.append((hub_rows[i][0], "codigo_erp_sem_preco_na_planilha"))
                continue
            preco = preco_por_codigo[codigo_erp_int]
            if not preco:
                pendencias.append((hub_rows[i][0], "preco_zero_ou_vazio_no_erp"))
                continue
            row = hub_rows[i][:]
            row[idx_preco] = round(float(preco), 2)
            conciliados.append(row)

    print(f"Conciliados: {len(conciliados)}")
    print(f"Pendencias: {len(pendencias)}")

    # 4. gravar xlsx final no mesmo formato do hub
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Produtos"
    ws_out.append(headers)
    for row in conciliados:
        ws_out.append(row)
    wb_out.save(OUT_XLSX)

    # 5. gravar pendencias
    with open(OUT_PENDENCIAS, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["nome", "motivo"])
        w.writerows(pendencias)

    print(f"\nOK -> {OUT_XLSX} ({len(conciliados)} linhas)")
    print(f"Pendencias -> {OUT_PENDENCIAS} ({len(pendencias)} linhas)")


if __name__ == "__main__":
    main()
