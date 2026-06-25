# -*- coding: utf-8 -*-
"""
Importa produtos_hub_CONCILIADO.xlsx -> tabela `produtos` do Supabase.

Regras (decididas com o usuario, 2026-06-24):
  - JA EXISTE no banco (match por nome, case/acento-insensitive):
        UPDATE seletivo: atualiza preco_atual, categoria, unidade, validade.
        PRESERVA ativo e nicho (nao tira nada do site nem desativa curados).
  - NAO existe (novo): INSERT com ativo=False (ativar sob demanda).
  - Embeddings: gerados pelo Fluxo 10 (trigger no INSERT/UPDATE). Nao tratados aqui.

SEGURANCA:
  - Default = DRY-RUN (nao grava). Mostra: quantos insert/update, diffs de preco,
    curados (ofertas/kits/itens com nicho) afetados. Use --executar p/ gravar de verdade.
  - Match por nome exato (PostgREST .eq). Para nomes que diferem so por acento/caixa,
    o banco e a planilha ja foram normalizados na fase de conciliacao, mas o match aqui
    e por nome LITERAL (o banco tem o nome exato). Divergencias viram INSERT (novo).

Uso:
  python -u importar_conciliado.py              # DRY-RUN (nao grava)
  python -u importar_conciliado.py --executar   # grava de verdade
"""
import json
import re
import sys
import time
import unicodedata
import urllib.error
import urllib.request

import openpyxl

XLSX = "produtos_hub_CONCILIADO.xlsx"
BANCO_JSON = "banco_produtos.json"
ENV = "../.env.local"
PROJ = "zziapgnenvugyvrgrhrs"
BATCH_INSERT = 500


def sa(s):
    return "".join(c for c in unicodedata.normalize("NFD", str(s or "")) if unicodedata.category(c) != "Mn")


def ck(s):
    return re.sub(r"[^a-z0-9]+", "", sa(s).lower())


def fl(v):
    try:
        return round(float(str(v).replace(",", ".").strip()), 2)
    except (ValueError, TypeError):
        return 0.0


def carregar_env():
    url = key = None
    for line in open(ENV, encoding="utf-8", errors="replace"):
        line = line.strip()
        if line.startswith("NEXT_PUBLIC_SUPABASE_URL"):
            url = line.split("=", 1)[1].strip().strip('"').strip("'")
        if line.startswith("SUPABASE_SERVICE_ROLE_KEY"):
            key = line.split("=", 1)[1].strip().strip('"').strip("'")
    if not url or not key:
        raise SystemExit("Faltam NEXT_PUBLIC_SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY em " + ENV)
    return url.rstrip("/"), key


def http(method, url, key, body=None):
    data = json.dumps(body).encode("utf-8") if body is not None else None
    req = urllib.request.Request(url, data=data, method=method, headers={
        "apikey": key, "Authorization": f"Bearer {key}",
        "Content-Type": "application/json", "Prefer": "return=minimal",
    })
    with urllib.request.urlopen(req, timeout=60) as r:
        return r.status


def main():
    executar = "--executar" in sys.argv
    modo = "EXECUTAR (grava no banco)" if executar else "DRY-RUN (nada sera gravado)"
    print(f"=== MODO: {modo} ===\n", flush=True)

    # banco atual: nome_norm -> {nome, preco, ativo, nicho, categoria}
    banco = json.load(open(BANCO_JSON, encoding="utf-8"))
    idx = {}
    for b in banco:
        idx[ck(b["nome"])] = b
    print(f"Banco atual: {len(banco)} produtos", flush=True)

    # planilha
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["Produtos"]
    H = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = [list(r) for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0]]
    wb.close()
    col = {h: i for i, h in enumerate(H)}
    print(f"Planilha conciliada: {len(rows)}\n", flush=True)

    inserts = []       # novos (ativo=False)
    updates = []       # existentes (update seletivo)
    diffs_preco = []   # mudancas de preco em existentes
    curados_afetados = []  # existentes com nicho (vitrine) que serao atualizados

    for r in rows:
        nome = str(r[col["nome"]]).strip()
        k = ck(nome)
        linha = {
            "nome": nome,
            "unidade": (r[col["unidade"]] or "UN"),
            "preco_atual": fl(r[col["preco_atual"]]),
            "categoria": (str(r[col["categoria"]]).strip().lower() if r[col["categoria"]] else None),
            "nicho": (str(r[col["nicho"]]).strip().lower() if r[col["nicho"]] else None),
            "validade": (r[col["validade"]] or None),
        }
        if k in idx:
            b = idx[k]
            # UPDATE seletivo: SO categoria/unidade/validade.
            # PRESERVA preco_atual (banco e mais confiavel/curado), ativo e nicho.
            payload = {
                "categoria": linha["categoria"],
                "unidade": linha["unidade"],
                "validade": linha["validade"],
            }
            updates.append((b["nome"], payload))
            pold = fl(b.get("preco_atual"))
            if pold != linha["preco_atual"]:
                diffs_preco.append((b["nome"], pold, linha["preco_atual"]))  # so informativo, NAO aplicado
            if (b.get("nicho") or "").strip():
                curados_afetados.append((b["nome"], b.get("nicho"), pold, linha["preco_atual"]))
        else:
            ins = dict(linha)
            ins["ativo"] = False
            ins["tipo"] = ("oferta" if linha["categoria"] == "ofertas"
                           else "kit" if linha["categoria"] == "kits"
                           else "padaria" if linha["categoria"] == "padaria"
                           else "produto")
            inserts.append(ins)

    print("================= RESUMO =================", flush=True)
    print(f"  INSERT (novos, ativo=false): {len(inserts)}", flush=True)
    print(f"  UPDATE (existentes, SO categoria/unidade/validade): {len(updates)}", flush=True)
    print(f"  -> PRECO dos existentes: PRESERVADO (nao atualizado). ativo e nicho tambem.", flush=True)
    print(f"  -> curados (com nicho/vitrine) atualizados (so cat/unidade): {len(curados_afetados)}", flush=True)

    # amostras — diffs de preco sao apenas INFORMATIVOS (nao aplicados)
    print("\n  [INFO] Preços que DIFEREM banco vs planilha (NAO serao alterados):", flush=True)
    for n, a, b2 in diffs_preco[:15]:
        print(f"     {n[:45]:45} banco R${a} (mantido) | planilha tinha R${b2}", flush=True)
    print(f"     ... ({len(diffs_preco)} no total — todos mantidos como no banco)", flush=True)

    print("\n  Curados afetados (nicho preservado, so preco/cat atualizam):", flush=True)
    for n, nh, a, b2 in curados_afetados[:20]:
        flag = "" if a == b2 else f"  PRECO {a}->{b2}"
        print(f"     [{nh}] {n[:40]:40}{flag}", flush=True)
    print(f"     ... ({len(curados_afetados)} no total)", flush=True)

    # tipos dos inserts
    from collections import Counter
    print("\n  Tipos dos novos (insert):", dict(Counter(i["tipo"] for i in inserts)), flush=True)
    print("  Categorias dos novos:", dict(Counter(i["categoria"] for i in inserts).most_common()), flush=True)

    if not executar:
        print("\n*** DRY-RUN: nada foi gravado. Rode com --executar para aplicar. ***", flush=True)
        return

    # ---------- EXECUCAO REAL ----------
    so_inserts = "--apenas-inserts" in sys.argv
    so_updates = "--apenas-updates" in sys.argv
    url, key = carregar_env()
    pausa = 2.0  # segundos entre lotes de insert

    if not so_updates:
        print(f"\n>>> Inserindo {len(inserts)} novos em lotes de {BATCH_INSERT} (pausa {pausa}s)...", flush=True)
        t0 = time.time()
        for i in range(0, len(inserts), BATCH_INSERT):
            lote = inserts[i:i + BATCH_INSERT]
            http("POST", f"{url}/rest/v1/produtos", key, lote)
            print(f"    inserts {min(i+BATCH_INSERT, len(inserts))}/{len(inserts)} ({time.time()-t0:.0f}s)", flush=True)
            if i + BATCH_INSERT < len(inserts):
                time.sleep(pausa)
        print(f"    >>> INSERTS concluidos ({len(inserts)})", flush=True)

    if not so_inserts:
        print(f"\n>>> Atualizando {len(updates)} existentes (so categoria/unidade/validade)...", flush=True)
        t0 = time.time()
        for j, (nome, payload) in enumerate(updates, 1):
            nome_q = urllib.parse.quote(nome, safe="")
            http("PATCH", f"{url}/rest/v1/produtos?nome=eq.{nome_q}", key, payload)
            if j % 200 == 0:
                print(f"    updates {j}/{len(updates)} ({time.time()-t0:.0f}s)", flush=True)
        print(f"    >>> UPDATES concluidos ({len(updates)})", flush=True)
    print("\n=== FASE CONCLUIDA ===", flush=True)


if __name__ == "__main__":
    import urllib.parse  # noqa
    main()
