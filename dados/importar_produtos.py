# -*- coding: utf-8 -*-
"""
Importa produtos_relatorio_original_formatação_padrão_hub.xlsx -> tabela `produtos` do Supabase.

- Todos entram com ativo=False e preco_atual=0 (precos serao definidos depois)
- Usa INSERT ... ON CONFLICT (nome) DO NOTHING  ->  nao sobrescreve os 421 existentes
- Envia em lotes de 500 via REST API do Supabase (prefer=resolution=ignore-duplicates)

Uso:
  python importar_produtos.py            # importa tudo
  python importar_produtos.py --dry-run  # mostra quantos seriam inseridos, sem tocar no banco
"""
import json, os, re, sys, time
import openpyxl
import urllib.request, urllib.error

XLSX = "produtos_relatorio_original_formatação_padrão_hub.xlsx"
ENV  = "../.env.local"
SUPABASE_PROJECT = "zziapgnenvugyvrgrhrs"
BATCH = 500

# ---- helpers ----
def carregar_env():
    url = key = None
    for line in open(ENV, encoding="utf-8", errors="replace"):
        line = line.strip()
        if line.startswith("NEXT_PUBLIC_SUPABASE_URL"):
            url = line.split("=", 1)[1].strip().strip('"').strip("'")
        if line.startswith("SUPABASE_SERVICE_ROLE_KEY"):
            key = line.split("=", 1)[1].strip().strip('"').strip("'")
    if not url or not key:
        raise SystemExit("NEXT_PUBLIC_SUPABASE_URL ou SUPABASE_SERVICE_ROLE_KEY nao encontrados em " + ENV)
    return url.rstrip("/"), key

def supabase_insert(url, key, rows, dry_run=False):
    """Insere lote via REST. ON CONFLICT (nome) DO NOTHING via prefer header."""
    endpoint = f"{url}/rest/v1/produtos?on_conflict=nome"
    body = json.dumps(rows).encode("utf-8")
    if dry_run:
        return len(rows)
    req = urllib.request.Request(
        endpoint,
        data=body,
        method="POST",
        headers={
            "apikey": key,
            "Authorization": f"Bearer {key}",
            "Content-Type": "application/json",
            "Prefer": "resolution=ignore-duplicates,return=minimal",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as r:
            return len(rows)
    except urllib.error.HTTPError as e:
        body_err = e.read().decode("utf-8", errors="replace")
        raise SystemExit(f"Erro HTTP {e.code}: {body_err}")

def limpar_preco(v):
    if v is None: return 0
    s = str(v).replace(",", ".").strip()
    try:
        f = float(s)
        return round(f, 2) if f > 0 else 0
    except ValueError:
        return 0

# ---- main ----
def main():
    dry_run = "--dry-run" in sys.argv
    if dry_run:
        print("[DRY-RUN] Nenhuma alteracao sera feita no banco.")

    url, key = carregar_env()

    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb["Produtos"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    print(f"Colunas: {headers}")

    col = {h: i for i, h in enumerate(headers)}

    rows_to_insert = []
    pulados = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        nome = row[col["nome"]] if "nome" in col else None
        if not nome or str(nome).strip() == "":
            pulados += 1
            continue

        preco = limpar_preco(row[col.get("preco_atual", -1)] if "preco_atual" in col else 0)
        unidade = str(row[col["unidade"]]).strip() if "unidade" in col and row[col["unidade"]] else "UN"
        categoria = str(row[col["categoria"]]).strip().lower() if "categoria" in col and row[col["categoria"]] else None
        nicho = str(row[col["nicho"]]).strip().lower() if "nicho" in col and row[col["nicho"]] else None
        validade = str(row[col["validade"]]).strip() if "validade" in col and row[col["validade"]] else None

        # limpar "None" strings
        if nicho in (None, "", "none"): nicho = None
        if validade in (None, "", "none"): validade = None
        if categoria in (None, "", "none", "a definir"): categoria = "outros"

        rows_to_insert.append({
            "nome": str(nome).strip(),
            "preco_atual": preco,
            "unidade": unidade,
            "categoria": categoria,
            "nicho": nicho,
            "validade": validade,
            "tipo": "produto",
            "ativo": False,   # todos inativos ate ter preco definido
        })

    wb.close()
    print(f"Lidos: {len(rows_to_insert)} | Pulados (vazios): {pulados}")

    if not rows_to_insert:
        print("Nada para inserir.")
        return

    # enviar em lotes
    total_inserido = 0
    lotes = [rows_to_insert[i:i+BATCH] for i in range(0, len(rows_to_insert), BATCH)]
    print(f"Enviando {len(lotes)} lotes de ate {BATCH}...")
    t0 = time.time()
    for i, lote in enumerate(lotes, 1):
        n = supabase_insert(url, key, lote, dry_run=dry_run)
        total_inserido += n
        el = time.time() - t0
        print(f"  lote {i}/{len(lotes)} | {total_inserido} enviados | {el:.0f}s", flush=True)

    print(f"\n{'[DRY-RUN] ' if dry_run else ''}Concluido: {total_inserido} produtos {'seriam ' if dry_run else ''}enviados ao Supabase.")
    print("Todos com ativo=False. Quando tiver precos, exporte, preencha e reimporte com script de update.")

if __name__ == "__main__":
    main()
