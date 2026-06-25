# -*- coding: utf-8 -*-
"""
Trata produtos_hub_precos_conciliados.xlsx -> formato final p/ o catalogo do Hub/Supabase.

Regras (definidas com o usuario, sessao 2026-06-24):

UNIDADE:
  - categoria=carnes SEM peso no nome (sem '500g','1kg','800g'...) -> 'Kg' (balanca)
  - carne COM peso no nome -> 'UN'
  - 'pacote' detectavel no nome -> 'pct'
  - resto -> 'UN'
  - unidades quebradas (lixo da descricao: '5L.', 'GOTAS SUAVE.') -> recalculadas pela regra acima

CATEGORIA (lowercase, sem acento):
  - mantem as descritivas reais (mercearia, bebidas, limpeza, higiene, laticinios,
    temperos, congelados, carnes, padaria, pet, outros)
  - 'ofertas' e 'kits' SAO PROIBIDOS aqui (exclusivos do catalogo atual do Hub) ->
    itens marcados assim + 'Frios'/'a definir' sao reclassificados p/ a categoria real
  - reclassifica 'outros' por palavras-chave determinísticas quando obvio

NICHO: tudo vazio (etapa posterior).

DEDUP (nome identico, case-insensitive):
  - manter 1: MAIOR preco; empate -> categoria mais especifica (!= 'outros')
  - lista os removidos em relatorio

NEAR-DUPLICATES: so LISTA (nunca remove por similaridade).

SAIDA: mesmo formato do Hub: nome, unidade, preco_atual, categoria, nicho, validade
  - NAO importa nada. Gera arquivo + relatorios p/ revisao manual.

Uso:
  python -u tratar_planilha_hub.py
"""
import csv
import re
import sys
import unicodedata
from collections import defaultdict

import openpyxl

ENTRADA = "produtos_hub_precos_conciliados.xlsx"
SAIDA_XLSX = "produtos_hub_TRATADO.xlsx"
REL_DEDUP = "rel_dedup_removidos.csv"
REL_NEARDUP = "rel_near_duplicates.csv"
REL_RECLASS = "rel_reclassificados.csv"
REL_CATEGORIAS = "rel_categorias_final.csv"

CATEGORIAS_VALIDAS = {
    "mercearia", "bebidas", "limpeza", "higiene", "laticinios",
    "temperos", "congelados", "carnes", "padaria", "pet", "outros",
}
# categoria proibida nos 27k (so vitrine curada do hub) -> reclassificar
CATEGORIA_PROIBIDA = {"ofertas", "oferta", "kits", "kit"}


def sem_acento(s):
    s = str(s or "")
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def norm_cat(c):
    return sem_acento(c).strip().lower()


# peso no nome: 500g, 1kg, 1,5kg, 800 g, 2l, 350ml, 90g, etc.
RE_PESO = re.compile(r"\d+\s*[.,]?\s*\d*\s*(kg|kgs|g|gr|gramas?|ml|l|lt|litros?|cl)\b", re.IGNORECASE)
# pacote/multipack
RE_PACOTE = re.compile(r"\b(pacote|pacotes|pct|fardo|fardos|c/\s*\d+\s*un|com\s+\d+\s+unidades?|leve\s+\d+)\b", re.IGNORECASE)
# unidades validas que ja vem certas
UNID_OK = {"un", "kg", "pct", "pacote", "unidade", "uni"}

# --- carnes: Kg SOMENTE para corte in natura + frios a peso (allowlist) ---
# Decisao do usuario (2026-06-24): na duvida UN. So vira Kg se bater nesta lista.
# Excecao explicita: steak/hamburguer/nuggets -> SEMPRE UN (mesmo de açougue).
CARNE_KG_KEYWORDS = [
    # bovino
    "alcatra", "acem", "costela", "costelinha", "picanha", "contra file", "contrafile",
    "coxao", "fraldinha", "musculo", "lombo", "patinho", "maminha", "cupim", "file mignon",
    "ponta de agulha", "rabada", "rabo", "lagarto", "largato", "peito com osso", "dianteiro",
    "traseiro", "osso buco", "bife", "chuleta", "chuletinha", "suan", "bucho", "fiambre bovino",
    "picadinho", "picadao", "miolo", "paleta", "pa bovina", "carne moida", "carne de sol",
    "carne seca", "fígado", "figado", "moela", "miudos", "miudezas", "coracao", "lingua",
    "orelha", "joelho", "joelho de porco", "pe de", "tripa", "toucinho", "carcaca", "carcaça",
    # suino
    "pernil", "bisteca", "copa lombo", "barriga suina", "filé suino", "file suino", "filé de pernil",
    "file de pernil", "porco", "suina", "suino", "leitao",
    # aves
    "coxa", "sobrecoxa", "coxinha da asa", "asa", "tulipa", "file de frango", "filé de frango",
    "peito de frango", "frango inteiro", "frango a passarinho", "frango passarinho",
    "frango desossado", "frango resfriado", "sassami", "passarinho",
    # peixe in natura
    "tilapia", "tilápia", "merluza", "pescada", "cacao", "cação", "posta", "sardinha fresca",
    "salmao", "salmão", "peixe",
    # frios/embutidos a peso (confirmado pelo usuario: ficam Kg)
    "bacon", "salsicha", "mortadela", "presunto", "calabresa", "linguica", "linguiça",
    "paio", "salame", "salsichao", "salsichão", "lombo canadense", "peito de peru",
    "peito de peru defumado", "blanquet", "apresuntado", "pastrami", "copa ",
    "toscana", "cuiabana", "cuaibana",  # tipos de linguica vendidos a peso
]
# excecoes: industrializados que viram UN mesmo sendo "carnes"
CARNE_UN_OVERRIDE = ["steak", "hamburguer", "hambúrguer", "nugget", "nuggets", "empanado",
                     "hot dog", "hotdog", "almondega", "almôndega", "bolinho", "kibe",
                     "quibe", "frango desfiado", "espetinho", "espeto", "beef burger"]


def tem_peso(nome):
    return bool(RE_PESO.search(nome or ""))


def eh_pacote(nome):
    return bool(RE_PACOTE.search(nome or ""))


def _carne_eh_kg(nome):
    """True se a carne deve ser Kg (corte in natura / frio a peso). Default = False (UN)."""
    n = " " + sem_acento(nome).lower() + " "
    # excecao tem prioridade: steak/hamburguer/nuggets/etc -> UN
    if any(sem_acento(k).lower() in n for k in CARNE_UN_OVERRIDE):
        return False
    return any(sem_acento(k).lower() in n for k in CARNE_KG_KEYWORDS)


def calcular_unidade(nome, categoria, unidade_orig):
    cat = norm_cat(categoria)
    if eh_pacote(nome):
        return "pct"
    # carnes: Kg so para corte in natura/frio a peso, sem peso explicito no nome
    if cat == "carnes" and not tem_peso(nome) and _carne_eh_kg(nome):
        return "Kg"
    # se a unidade original ja era um peso valido (kg) preservar
    uo = str(unidade_orig or "").strip().lower()
    if uo == "kg":
        return "Kg"
    return "UN"


# --- Reclassificacao de 'outros' (sessao 2026-06-24, ENDURECIDA) ---
# Licao: keyword-match cego como substring gerou ~100 falsos positivos
# (agua->bebidas pegava 'agua oxigenada'; coracao->pet pegava 'vela Sagrado Coracao';
# oleo->temperos pegava 'inseticida com oleo'; leite->laticinios pegava 'Leite de Rosas';
# bolo->padaria pegava 'bandeja para bolo'). Correcoes:
#  1) match por PALAVRA INTEIRA (regex \b), nao substring;
#  2) so keywords de ALTA confianca (sem agua/oleo/leite/coracao/bolo/rosca/cha soltas);
#  3) lista de EXCLUSAO: nome com termo de nao-alimento/utilidade -> fica em 'outros'.
# Na duvida, NAO reclassifica (deixa 'outros' — categoria valida).

# se o nome contiver qualquer um destes, NUNCA reclassifica (fica em outros)
EXCLUI_RECLASS = [
    "esmalte", "sandalia", "havaianas", "vela", "lampada", "inseticida", "raid", "baygon",
    "papel", "bandeja", "tabua", "luva", "bucha", "resistencia", "chuveiro", "ducha",
    "borracha", "balao", "forminha", "forma de", "espatula", "bico", "faca", "pa de lixo",
    "prato", "franja", "niple", "abracadeira", "fita", "rosca", "agua oxigenada",
    "leite de rosas", "manteiga de cacau", "papel manteiga", "oleo capilar", "oleo lubrificante",
    "oleo singer", "oleo para maquina", "oleo de amendoa", "oleo de argan", "agua para bateria",
    "agua raz", "creme de tratamento", "convite", "saboneteira", "recipiente", "concha",
    "aparelho eletronico", "desodorante ambiente", "desodorante de ambiente",
    # utensilios/itens que confundiam congelados/bebidas/limpeza:
    "copo", "caneca", "palito", "palitos", "difusor", "aromatizador", "aromatizante",
    "esponja para banho", "esponja de banho", "saboneteira", "alho congelado",
]

# keywords por categoria — match por PALAVRA INTEIRA. So termos sem ambiguidade.
REGRAS_CAT = [
    ("congelados", ["sorvete", "picole", "pizza", "lasanha", "nuggets", "congelado", "congelada",
                    "acai", "polpa de fruta"]),
    ("limpeza", ["sabao", "detergente", "desinfetante", "amaciante", "alvejante", "vassoura",
                 "rodo", "esponja", "saco de lixo", "agua sanitaria", "lustra", "limpador",
                 "lava roupa", "lava louca", "cloro", "agua sanitaria"]),
    ("higiene", ["shampoo", "sabonete", "condicionador", "absorvente", "fralda", "cotonete",
                 "creme dental", "escova de dente", "papel higienico", "antisseptico bucal",
                 "fio dental", "hidratante corporal", "desodorante roll", "desodorante aerosol"]),
    ("pet", ["racao", "petisco para caes", "petisco para gatos", "areia higienica",
             "sache para gato", "sache para cachorro", "osso para cachorro"]),
    ("laticinios", ["iogurte", "queijo", "mussarela", "mucarela", "requeijao", "margarina",
                    "manteiga", "petit suisse", "leite condensado", "creme de leite",
                    "leite fermentado", "leite integral", "leite desnatado", "danone"]),
    ("temperos", ["tempero", "oregano", "colorau", "acafrao", "vinagre", "caldo knorr",
                  "caldo de", "molho de pimenta", "molho de churrasco", "sal grosso",
                  "sal refinado", "pimenta do reino", "azeite de oliva"]),
    ("padaria", ["pao frances", "pao de forma", "pao de hamburguer", "pao de queijo", "broa",
                 "rosca doce", "panettone", "panetone", "bisnaguinha"]),
    ("carnes", ["linguica", "mortadela", "presunto", "salsicha fresca", "bacon", "salame",
                "frango resfriado", "carne moida", "costela bovina", "alcatra", "picanha",
                "contra file", "coxa e sobrecoxa"]),
    ("bebidas", ["refrigerante", "cerveja", "suco", "energetico", "vodka", "whisky", "cachaca",
                 "agua mineral", "agua tonica", "refresco em po", "isotonico", "vinho tinto",
                 "vinho branco", "espumante", "chopp"]),
]

# compilar regex por palavra inteira (sem acento, lower)
_REGRAS_RE = [(cat, [re.compile(r"(?<![a-z0-9])" + re.escape(sem_acento(k).lower()) + r"(?![a-z0-9])")
                     for k in kws]) for cat, kws in REGRAS_CAT]
_EXCLUI_RE = [re.compile(r"(?<![a-z0-9])" + re.escape(sem_acento(k).lower()) + r"(?![a-z0-9])")
              for k in EXCLUI_RECLASS]


def reclassificar(nome, cat_atual):
    """Retorna (categoria_final, motivo|None). So mexe se invalida/proibida/'outros'."""
    catn = norm_cat(cat_atual)
    if catn in CATEGORIA_PROIBIDA:
        nova = _por_palavra(nome) or "outros"
        return nova, f"proibida({catn})->{nova}"
    if catn in ("", "a definir", "frios", "(vazio)", "none"):
        nova = _por_palavra(nome) or "outros"
        return nova, f"invalida({catn})->{nova}"
    if catn == "outros":
        nova = _por_palavra(nome)
        if nova:
            return nova, f"outros->{nova}"
        return "outros", None
    if catn in CATEGORIAS_VALIDAS:
        return catn, None
    # categoria desconhecida -> tenta palavra, senao outros
    nova = _por_palavra(nome) or "outros"
    return nova, f"desconhecida({catn})->{nova}"


def _por_palavra(nome):
    n = sem_acento(nome).lower()
    # exclusao: nao-alimento/utilidade -> nao reclassifica
    if any(rx.search(n) for rx in _EXCLUI_RE):
        return None
    for cat, regexes in _REGRAS_RE:
        if any(rx.search(n) for rx in regexes):
            return cat
    return None


def limpar_preco(v):
    if v is None:
        return 0.0
    try:
        return round(float(str(v).replace(",", ".").strip()), 2)
    except ValueError:
        return 0.0


def especificidade(cat):
    # categoria mais especifica vence empate de preco; 'outros' e a menos especifica
    return 0 if norm_cat(cat) == "outros" else 1


# --- overrides de dedup decididos com o usuario (2026-06-24), por nome (lower) ---
# 'remover' = descartar o grupo inteiro; 'menor' = manter o de menor preco; 'maior' = manter maior.
# Regra padrao dos demais grupos = manter MAIOR preco.
DEDUP_OVERRIDE = {
    # remover do catalogo
    "brigadeirao grande": "remover",
    "brigadeirão grande": "remover",
    "cafe brasileiro 500g": "remover",
    "café brasileiro 500g": "remover",
    # bebidas lata/garrafa pequena: preco maior = fardo vazado -> manter unitario (menor)
    "refrigerante coca-cola 250ml pet": "menor",
    "refrigerante fanta uva zero 350ml": "menor",
    "cerveja nova schin 350ml": "menor",
    "refrigerante sukita laranja 350ml": "menor",
    "coco baianinha 900ml": "menor",
    "vodka kadov ice 275ml": "menor",
    "cerveja skol beats 269ml": "menor",
    # tamanho: versao grande/kg ja existe como item proprio -> manter unitario (menor)
    "torta holandesa pequena": "menor",
    "mini sonho": "menor",
    # OBS: os 6 ambiguos sem evidencia (Sabao em Po Brilhante, Pudim de Leite Condensado,
    # Baguete de Gergelim, Esfiha, Vassoura Varre Bem, Amido de Milho Maizena) seguem a
    # regra PADRAO = maior preco. Nao precisam estar aqui.
}


# --- fusao de near-duplicates (mesmo produto, grafia diferente) decidida com usuario ---
# chave = nome sem acento/pontuacao/espaco (chave_near). valor = (nome_final_limpo, preco_final).
# 8 grupos gap<=10% -> maior preco + nome limpo (automatico). 17 grupos gap>10% -> escolha do usuario.
# Os 'preco_final': None = manter o MAIOR do grupo automaticamente (usado nos 8 de gap pequeno).
NEARDUP_MERGE = {
    # --- 17 com gap >10% (nome limpo + preco escolhido pelo usuario) ---
    "bandejalaminadan7": ("Bandeja Laminada N°7", 4.25),
    "aparelhodebarbeargillettemach3": ("Aparelho de Barbear Gillette Mach 3", 19.99),
    "sucosufreshmorango200ml": ("Suco Sufresh Morango 200ml", 2.75),
    "sabaoempobrilhantelimpezatotal800g": ("Sabão em Pó Brilhante Limpeza Total 800g", 12.99),
    "saboneteprotex90gervadoce": ("Sabonete Protex 90g Erva-doce", 2.25),
    "cervejaschinmalzbier355ml": ("Cerveja Schin Malzbier 355ml", 2.75),
    "alvejantevanish450gpo2crystalwhite": ("Alvejante Vanish 450g Pó 2 Crystal White", 19.99),
    "cafe3coracoesextraforte500g": ("Café 3 Corações Extra Forte 500g", 27.99),
    "sabaoemposapolioradium300g": ("Sabão em Pó Sapólio Radium 300g", 7.99),
    "papelhigienicobianco60m": ("Papel Higiênico Bianco 60m", 14.99),
    "esponjascotchbritecom3unidades": ("Esponja Scotch-Brite com 3 Unidades", 4.99),
    "azeitegallo500ml": ("Azeite Gallo 500ml", 29.99),
    "bandejalaminadan07": ("Bandeja Laminada N°07", 4.50),
    "bandejalaminadan05": ("Bandeja Laminada N°05", 3.00),
    "farinhalacteanutrifoods240g": ("Farinha Láctea Nutrifoods 240g", 3.79),
    "cigarrorothmansclicksensemelao": ("Cigarro Rothmans Click Sense Melão", 9.75),
    "cafebrasileiroextraforte500g": ("Café Brasileiro Extra Forte 500g", 31.99),
    # --- 8 com gap <=10% (nome limpo, preco = MAIOR automatico) ---
    "sucodafrutalaranja1l": ("Suco da Fruta Laranja 1L", None),
    "queijoraladoreliquiadacanastra50g": ("Queijo Ralado Relíquia da Canastra 50g", None),
    "ovodepascoalecacau100g": ("Ovo de Páscoa Lecacau 100g", None),
    "rododeplastico30cm": ("Rodo de Plástico 30cm", None),
    "sabaoempoblanc1kg": ("Sabão em Pó Blanc 1kg", None),
    "docedeamendoimgideao130g": ("Doce de Amendoim Gideão 130g", None),
    "bandejalaminadan6": ("Bandeja Laminada N°6", None),
    "leitequatasemidesnatado1l": ("Leite Quatá Semidesnatado 1L", None),
}


def main():
    print(f"Lendo {ENTRADA}...", flush=True)
    wb = openpyxl.load_workbook(ENTRADA, read_only=True, data_only=True)
    ws = wb["Produtos"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    print(f"Header: {headers}", flush=True)
    col = {h: i for i, h in enumerate(headers)}

    brutos = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        nome = r[col["nome"]]
        if not nome or str(nome).strip() == "":
            continue
        brutos.append({
            "nome": str(nome).strip(),
            "unidade_orig": r[col["unidade"]],
            "preco": limpar_preco(r[col["preco_atual"]]),
            "categoria_orig": r[col["categoria"]],
            "validade": r[col.get("validade", -1)] if "validade" in col else None,
        })
    wb.close()
    print(f"Lidos: {len(brutos)}", flush=True)

    # ---- transformar: unidade + categoria ----
    reclass_log = []
    for b in brutos:
        b["unidade"] = calcular_unidade(b["nome"], b["categoria_orig"], b["unidade_orig"])
        nova_cat, motivo = reclassificar(b["nome"], b["categoria_orig"])
        b["categoria"] = nova_cat
        if motivo:
            reclass_log.append((b["nome"], b["categoria_orig"], nova_cat, motivo))

    # ---- dedup por nome (case-insensitive) ----
    grupos = defaultdict(list)
    for b in brutos:
        grupos[b["nome"].strip().lower()].append(b)

    finais = []
    dedup_log = []
    removidos_grupo_log = []  # grupos inteiros descartados por override 'remover'
    for k, itens in grupos.items():
        override = DEDUP_OVERRIDE.get(k)
        if override == "remover":
            for it in itens:
                removidos_grupo_log.append((it["nome"], it["preco"], it["categoria"]))
            continue
        if len(itens) == 1:
            finais.append(itens[0])
            continue
        # ordenacao base: MAIOR preco; empate -> categoria mais especifica
        itens_ord = sorted(itens, key=lambda x: (x["preco"], especificidade(x["categoria"])), reverse=True)
        if override == "menor":
            # manter o de MENOR preco (override decidido com evidencia)
            itens_ord = sorted(itens, key=lambda x: (x["preco"], -especificidade(x["categoria"])))
        mantido = itens_ord[0]
        finais.append(mantido)
        for rem in itens_ord[1:]:
            dedup_log.append((rem["nome"], rem["preco"], rem["categoria"],
                              mantido["preco"], mantido["categoria"]))

    print(f"Apos dedup: {len(finais)} (removidos por dedup {len(dedup_log)}, grupos descartados {len(removidos_grupo_log)})", flush=True)

    # ---- near-duplicates: fundir os decididos, listar o resto ----
    # chave = nome sem acento/pontuacao/espaco; agrupa nomes distintos que colidem
    def chave_near(s):
        return re.sub(r"[^a-z0-9]+", "", sem_acento(s).lower())

    near_idx = defaultdict(list)  # chave -> lista de itens finais (objetos)
    for b in finais:
        near_idx[chave_near(b["nome"])].append(b)

    near_log = []           # grupos NAO fundidos (so listados)
    neardup_merge_log = []  # grupos fundidos (auditoria)
    descartar_ids = set()   # id(obj) a remover de finais (cópias fundidas)

    for k, itens in near_idx.items():
        nomes_distintos = sorted({b["nome"] for b in itens})
        if len(nomes_distintos) <= 1:
            continue
        if k in NEARDUP_MERGE:
            nome_final, preco_final = NEARDUP_MERGE[k]
            if preco_final is None:
                preco_final = max(b["preco"] for b in itens)  # gap pequeno -> maior
            # escolhe um item-base (primeiro) p/ herdar unidade/categoria/validade
            base = sorted(itens, key=lambda x: -x["preco"])[0]
            base["nome"] = nome_final
            base["preco"] = preco_final
            for b in itens:
                if b is not base:
                    descartar_ids.add(id(b))
            neardup_merge_log.append((nomes_distintos, nome_final, preco_final))
        else:
            near_log.append(nomes_distintos)

    if descartar_ids:
        finais = [b for b in finais if id(b) not in descartar_ids]

    # ---- gravar saida ----
    wb_out = openpyxl.Workbook()
    wso = wb_out.active
    wso.title = "Produtos"
    wso.append(["nome", "unidade", "preco_atual", "categoria", "nicho", "validade"])
    for b in sorted(finais, key=lambda x: (x["categoria"], x["nome"])):
        wso.append([b["nome"], b["unidade"], b["preco"], b["categoria"], None,
                    b["validade"] if b["validade"] not in (None, "", "none") else None])
    wb_out.save(SAIDA_XLSX)
    print(f"OK -> {SAIDA_XLSX} ({len(finais)} linhas)", flush=True)

    # ---- relatorios ----
    with open(REL_DEDUP, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["nome_removido", "preco_removido", "cat_removido", "preco_mantido", "cat_mantido"])
        w.writerows(dedup_log)

    with open("rel_grupos_descartados.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["nome_descartado", "preco", "categoria", "motivo"])
        for n, p, c in removidos_grupo_log:
            w.writerow([n, p, c, "override_remover"])

    with open(REL_NEARDUP, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["nomes_parecidos (revisar manualmente)"])
        for grp in near_log:
            w.writerow([" || ".join(grp)])

    with open("rel_neardup_fundidos.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["nomes_originais", "nome_final", "preco_final"])
        for orig, nf, pf in neardup_merge_log:
            w.writerow([" || ".join(orig), nf, pf])

    with open(REL_RECLASS, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["nome", "categoria_original", "categoria_nova", "motivo"])
        w.writerows(reclass_log)

    # distribuicao final de categorias e unidades
    from collections import Counter
    catf = Counter(b["categoria"] for b in finais)
    unif = Counter(b["unidade"] for b in finais)
    with open(REL_CATEGORIAS, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["tipo", "valor", "qtd"])
        for c, n in catf.most_common():
            w.writerow(["categoria", c, n])
        for u, n in unif.most_common():
            w.writerow(["unidade", u, n])

    print("\n=== RESUMO ===", flush=True)
    print("Categorias finais:", dict(catf.most_common()), flush=True)
    print("Unidades finais:", dict(unif.most_common()), flush=True)
    print(f"Reclassificados: {len(reclass_log)} (ver {REL_RECLASS})", flush=True)
    print(f"Dedup removidos: {len(dedup_log)} (ver {REL_DEDUP})", flush=True)
    print(f"Near-duplicate FUNDIDOS: {len(neardup_merge_log)} (ver rel_neardup_fundidos.csv)", flush=True)
    print(f"Near-duplicate p/ revisar (restantes): {len(near_log)} (ver {REL_NEARDUP})", flush=True)


if __name__ == "__main__":
    main()
